import json
import os
import pytest
from openpyxl import load_workbook

from chunk_norris.evaluator.report import Report
from chunk_norris.chunkers.fixed import FixedChunker


# ── Helpers ───────────────────────────────────────────────────────────────────

def make_scores(
    best_combined: float = 0.8,
    best_recall: float = 0.9,
    best_bert: float = 0.7,
    n_relevant: int = 2,
    recall_threshold: float = 0.75,
) -> dict:
    return {
        "bert_scores":       [best_bert],
        "token_recalls":     [best_recall],
        "combined_scores":   [best_combined],
        "best_bert":         best_bert,
        "best_recall":       best_recall,
        "best_combined":     best_combined,
        "relevant_combined": [best_combined],
        "avg_relevant":      best_combined,
        "n_relevant":        n_relevant,
        "n_retrieved":       3,
        "recall_threshold":  recall_threshold,
    }


def make_result(
    question: str = "What is X?",
    expected_answer: str = "X is Y.",
    scores: dict | None = None,
) -> dict:
    return {
        "question":         question,
        "expected_answer":  expected_answer,
        "retrieved_chunks": [{"text": "chunk text", "metadata": {}}],
        "scores":           scores or make_scores(),
    }


def make_experiment(
    label: str,
    results: list[dict],
    chunker: object | None = None,
) -> dict:
    return {
        "chunker":        label,
        "chunker_object": chunker or FixedChunker(chunk_size=256, overlap=0.1),
        "results":        results,
    }


# ── TestReportInit ────────────────────────────────────────────────────────────

class TestReportInit:

    def test_empty_experiments_raises(self):
        with pytest.raises(ValueError, match="experiments"):
            Report(experiments=[])

    def test_single_experiment_accepted(self):
        exp = make_experiment("ChunkerA", [make_result()])
        report = Report(experiments=[exp])
        assert report is not None

    def test_multiple_experiments_accepted(self):
        experiments = [
            make_experiment("ChunkerA", [make_result()]),
            make_experiment("ChunkerB", [make_result()]),
        ]
        report = Report(experiments=experiments)
        assert report is not None

    def test_summary_computed_on_init(self):
        exp = make_experiment("ChunkerA", [make_result()])
        report = Report(experiments=[exp])
        assert report._summary is not None
        assert len(report._summary) == 1


# ── TestComputeSummary ────────────────────────────────────────────────────────

class TestComputeSummary:

    def test_summary_has_one_entry_per_experiment(self):
        experiments = [
            make_experiment("A", [make_result()]),
            make_experiment("B", [make_result()]),
            make_experiment("C", [make_result()]),
        ]
        report = Report(experiments=experiments)
        assert len(report._summary) == 3

    def test_summary_sorted_best_first(self):
        experiments = [
            make_experiment("Low",  [make_result(scores=make_scores(best_combined=0.3))]),
            make_experiment("High", [make_result(scores=make_scores(best_combined=0.9))]),
            make_experiment("Mid",  [make_result(scores=make_scores(best_combined=0.6))]),
        ]
        report = Report(experiments=experiments)
        scores = [e["avg_best_combined"] for e in report._summary]
        assert scores == sorted(scores, reverse=True)

    def test_summary_avg_best_combined_correct(self):
        results = [
            make_result(scores=make_scores(best_combined=0.8)),
            make_result(scores=make_scores(best_combined=0.6)),
        ]
        exp = make_experiment("A", results)
        report = Report(experiments=[exp])
        assert report._summary[0]["avg_best_combined"] == 0.7

    def test_summary_avg_best_recall_correct(self):
        results = [
            make_result(scores=make_scores(best_recall=1.0)),
            make_result(scores=make_scores(best_recall=0.5)),
        ]
        exp = make_experiment("A", results)
        report = Report(experiments=[exp])
        assert report._summary[0]["avg_best_recall"] == 0.75

    def test_summary_avg_best_bert_correct(self):
        results = [
            make_result(scores=make_scores(best_bert=0.8)),
            make_result(scores=make_scores(best_bert=0.4)),
        ]
        exp = make_experiment("A", results)
        report = Report(experiments=[exp])
        assert report._summary[0]["avg_best_bert"] == 0.6

    def test_summary_avg_overall_equals_avg_best_combined(self):
        exp = make_experiment("A", [make_result(scores=make_scores(best_combined=0.77))])
        report = Report(experiments=[exp])
        entry = report._summary[0]
        assert entry["avg_overall"] == entry["avg_best_combined"]

    def test_summary_contains_recall_threshold(self):
        exp = make_experiment("A", [make_result(scores=make_scores(recall_threshold=0.8))])
        report = Report(experiments=[exp])
        assert report._summary[0]["recall_threshold"] == 0.8

    def test_summary_contains_chunker_label(self):
        exp = make_experiment("MyChunker", [make_result()])
        report = Report(experiments=[exp])
        assert report._summary[0]["chunker"] == "MyChunker"


# ── TestBest ──────────────────────────────────────────────────────────────────

class TestBest:

    def test_returns_highest_avg_best_combined(self):
        experiments = [
            make_experiment("Low",  [make_result(scores=make_scores(best_combined=0.3))]),
            make_experiment("High", [make_result(scores=make_scores(best_combined=0.9))]),
        ]
        report = Report(experiments=experiments)
        assert report.best()["chunker"] == "High"

    def test_tiebreak_by_recall(self):
        # Same combined score — higher recall wins
        experiments = [
            make_experiment("LowRecall",  [make_result(scores=make_scores(
                best_combined=0.7, best_recall=0.6))]),
            make_experiment("HighRecall", [make_result(scores=make_scores(
                best_combined=0.7, best_recall=0.9))]),
        ]
        report = Report(experiments=experiments)
        assert report.best()["chunker"] == "HighRecall"

    def test_returns_dict(self):
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        assert isinstance(report.best(), dict)

    def test_best_contains_chunker_key(self):
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        assert "chunker" in report.best()


# ── TestBestChunker ───────────────────────────────────────────────────────────

class TestBestChunker:

    def test_returns_chunker_object(self):
        chunker = FixedChunker(chunk_size=256, overlap=0.1)
        exp = make_experiment("A", [make_result()], chunker=chunker)
        report = Report(experiments=[exp])
        result = report.best_chunker()
        assert result is chunker

    def test_returns_best_chunker_not_any_chunker(self):
        chunker_low  = FixedChunker(chunk_size=128, overlap=0.1)
        chunker_high = FixedChunker(chunk_size=512, overlap=0.1)
        experiments = [
            make_experiment("Low",  [make_result(scores=make_scores(best_combined=0.3))],
                           chunker=chunker_low),
            make_experiment("High", [make_result(scores=make_scores(best_combined=0.9))],
                           chunker=chunker_high),
        ]
        report = Report(experiments=experiments)
        assert report.best_chunker() is chunker_high

    def test_best_chunker_can_chunk_text(self):
        chunker = FixedChunker(chunk_size=64, overlap=0.0)
        exp = make_experiment("A", [make_result()], chunker=chunker)
        report = Report(experiments=[exp])
        best = report.best_chunker()
        chunks = best.chunk("Some document text to chunk here.")
        assert isinstance(chunks, list)
        assert len(chunks) > 0


# ── TestCompare ───────────────────────────────────────────────────────────────

class TestCompare:

    def test_compare_does_not_raise_single_experiment(self, capsys):
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        report.compare()  # should not raise

    def test_compare_does_not_raise_multiple_experiments(self, capsys):
        experiments = [
            make_experiment("A", [make_result(scores=make_scores(best_combined=0.8))]),
            make_experiment("B", [make_result(scores=make_scores(best_combined=0.5))]),
        ]
        report = Report(experiments=experiments)
        report.compare()  # should not raise

    def test_compare_prints_best_marker(self, capsys):
        experiments = [
            make_experiment("Loser", [make_result(scores=make_scores(best_combined=0.3))]),
            make_experiment("Winner", [make_result(scores=make_scores(best_combined=0.9))]),
        ]
        report = Report(experiments=experiments)
        report.compare()
        output = capsys.readouterr().out
        assert "<-- best" in output
        assert "Winner" in output


# ── TestToJson ────────────────────────────────────────────────────────────────

class TestToJson:

    def test_creates_file(self, tmp_path):
        path = str(tmp_path / "results.json")
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        report.to_json(path)
        assert os.path.exists(path)

    def test_valid_json(self, tmp_path):
        path = str(tmp_path / "results.json")
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        report.to_json(path)
        with open(path) as f:
            data = json.load(f)
        assert isinstance(data, dict)

    def test_contains_summary_key(self, tmp_path):
        path = str(tmp_path / "results.json")
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        report.to_json(path)
        with open(path) as f:
            data = json.load(f)
        assert "summary" in data

    def test_contains_experiments_key(self, tmp_path):
        path = str(tmp_path / "results.json")
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        report.to_json(path)
        with open(path) as f:
            data = json.load(f)
        assert "experiments" in data

    def test_chunker_object_not_in_json(self, tmp_path):
        path = str(tmp_path / "results.json")
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        report.to_json(path)
        # Should not raise TypeError for non-serialisable chunker object
        with open(path) as f:
            content = f.read()
        assert "FixedChunker" not in content or "chunk_size" not in content

    def test_summary_correct_number_of_entries(self, tmp_path):
        path = str(tmp_path / "results.json")
        experiments = [
            make_experiment("A", [make_result()]),
            make_experiment("B", [make_result()]),
        ]
        report = Report(experiments=experiments)
        report.to_json(path)
        with open(path) as f:
            data = json.load(f)
        assert len(data["summary"]) == 2


# ── TestToExcel ───────────────────────────────────────────────────────────────

class TestToExcel:

    def test_creates_file(self, tmp_path):
        path = str(tmp_path / "results.xlsx")
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        report.to_excel(path)
        assert os.path.exists(path)

    def test_has_summary_sheet(self, tmp_path):
        path = str(tmp_path / "results.xlsx")
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        report.to_excel(path)
        wb = load_workbook(path)
        assert "Summary" in wb.sheetnames

    def test_has_one_sheet_per_chunker(self, tmp_path):
        path = str(tmp_path / "results.xlsx")
        experiments = [
            make_experiment("ChunkerA", [make_result()]),
            make_experiment("ChunkerB", [make_result()]),
        ]
        report = Report(experiments=experiments)
        report.to_excel(path)
        wb = load_workbook(path)
        # Summary + one Config sheet per chunker
        assert len(wb.sheetnames) == 3
        assert "Summary" in wb.sheetnames
        assert "Config 1" in wb.sheetnames
        assert "Config 2" in wb.sheetnames

    def test_summary_sheet_has_correct_row_count(self, tmp_path):
        path = str(tmp_path / "results.xlsx")
        experiments = [
            make_experiment("A", [make_result()]),
            make_experiment("B", [make_result()]),
            make_experiment("C", [make_result()]),
        ]
        report = Report(experiments=experiments)
        report.to_excel(path)
        wb = load_workbook(path)
        ws = wb["Summary"]
        # Header row + one row per chunker
        assert ws.max_row == 4

    def test_summary_sheet_first_row_is_header(self, tmp_path):
        path = str(tmp_path / "results.xlsx")
        exp = make_experiment("A", [make_result()])
        report = Report(experiments=[exp])
        report.to_excel(path)
        wb = load_workbook(path)
        ws = wb["Summary"]
        # First column is now Config, second is Chunker
        assert ws.cell(row=1, column=1).value == "Config"
        assert ws.cell(row=1, column=2).value == "Chunker"

    def test_detail_sheet_has_correct_row_count(self, tmp_path):
        path = str(tmp_path / "results.xlsx")
        results = [make_result(question=f"Q{i}") for i in range(5)]
        exp = make_experiment("ChunkerA", results)
        report = Report(experiments=[exp])
        report.to_excel(path)
        wb = load_workbook(path)
        # Detail sheets are now named Config N, not by chunker label
        ws = wb["Config 1"]
        # Title row + header row + one row per question
        assert ws.max_row == 7
