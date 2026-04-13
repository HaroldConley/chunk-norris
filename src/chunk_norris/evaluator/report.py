import json
from typing import Any

from openpyxl import Workbook  # openpyxl==3.1.5
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Style constants ───────────────────────────────────────────────────────────

_HEADER_FILL   = PatternFill("solid", fgColor="2E4057")
_BEST_FILL     = PatternFill("solid", fgColor="D4EDDA")
_ALT_FILL      = PatternFill("solid", fgColor="F8F9FA")
_RELEVANT_FILL = PatternFill("solid", fgColor="FFF3CD")
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
_BOLD_FONT     = Font(bold=True, size=10)
_NORMAL_FONT   = Font(size=10)
_WRAP          = Alignment(wrap_text=True, vertical="top")
_CENTER        = Alignment(horizontal="center", vertical="top")
_THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


class Report:
    """
    Compares scored results across multiple chunker configurations and
    produces both a printed summary and an Excel workbook.

    Primary metric: avg_best_combined — the average of the best combined
    score (token recall + bert score) / 2 across all questions.

    The Excel workbook contains:
        - Sheet 1 "Summary": one row per chunker, averages, best highlighted
        - One sheet per chunker: full per-question breakdown with all scores
          and retrieved chunk texts

    Args:
        experiments (list[dict]): One entry per chunker configuration, each with:
            - "chunker" (str): A human-readable label for the chunker config.
            - "results" (list[dict]): Scored results from Metrics.score().

    Example::

        report.compare()
        best = report.best()
        report.to_excel("results.xlsx")
        report.to_json("results.json")
    """

    def __init__(self, experiments: list[dict[str, Any]]) -> None:
        if not experiments:
            raise ValueError("experiments must not be empty.")

        self.experiments = experiments
        self._summary = self._compute_summary()

    def compare(self) -> None:
        """
        Prints a formatted comparison table. Primary metric is Combined
        (avg of token recall + bert score). Best configuration is marked.

        Example output::

            Chunker                              Combined  Recall  Bert   NRel
            ---                                  --------  ------  ----   ----
            FixedChunker(chunk_size=128, ...)      0.8200  0.9100  0.730   2.3  <-- best
            FixedChunker(chunk_size=256, ...)      0.7400  0.8500  0.630   1.8
            FixedChunker(chunk_size=512, ...)      0.6200  0.7100  0.530   1.1
        """
        best_label = self.best()["chunker"]
        col_w = max(len(e["chunker"]) for e in self._summary)
        col_w = max(col_w, len("Chunker"))

        header = (
            f"{'Chunker':<{col_w}}  "
            f"{'Combined':>8}  "
            f"{'Recall':>6}  "
            f"{'Bert':>6}  "
            f"{'NRel':>4}"
        )
        print(header)
        print("-" * len(header))

        for entry in self._summary:
            marker = "  <-- best" if entry["chunker"] == best_label else ""
            print(
                f"{entry['chunker']:<{col_w}}  "
                f"{entry['avg_best_combined']:>8.4f}  "
                f"{entry['avg_best_recall']:>6.4f}  "
                f"{entry['avg_best_bert']:>6.4f}  "
                f"{entry['avg_n_relevant']:>4.1f}"
                f"{marker}"
            )

    def best(self) -> dict[str, Any]:
        """
        Returns the summary entry for the best configuration.

        Tie-breaking: higher avg_best_recall wins — completeness is preferred
        over focus when combined scores are equal.

        Returns a summary dict with scores. To get the actual chunker object
        ready to use in a pipeline, call best_chunker() instead.
        """
        return max(
            self._summary,
            key=lambda e: (e["avg_best_combined"], e["avg_best_recall"]),
        )

    def best_chunker(self) -> Any:
        """
        Returns the best performing chunker instance, ready to use in a pipeline.

        This is the primary way to integrate chunk-norris into a larger system.
        After running an evaluation, call this method to get the winning chunker
        and pass it directly to your RAG pipeline or vector store.

        Returns:
            BaseChunker: The chunker instance that produced the best scores.
                         Call .chunk(text) on it to produce chunks.

        Example::

            norris = Norris(embedder=BertEmbedder())
            report = norris.run(text=TEXT, chunkers=[...], questions=QUESTIONS)

            # Get the best chunker and use it directly
            best = report.best_chunker()
            chunks = best.chunk(TEXT)

            # Pass chunks to your RAG pipeline
            vector_store.add(chunks)
            rag_pipeline.index(chunks)

        Note:
            The returned chunker is the same instance used during evaluation.
            Do not modify its parameters after receiving it — create a new
            instance if you need different settings.
        """
        best_label = self.best()["chunker"]
        for experiment in self.experiments:
            if experiment["chunker"] == best_label:
                return experiment["chunker_object"]
        raise RuntimeError(
            f"Could not find chunker object for label: {best_label}"
        )

    def to_excel(self, path: str) -> None:
        """
        Exports results to an Excel workbook with multiple sheets.

        Sheet 1 — Summary: one row per chunker, averages, best highlighted.
        Sheet 2..N — one sheet per chunker with full per-question detail:
            - Question and expected answer
            - Best combined, recall, and bert scores
            - Each retrieved chunk with its individual scores
            - Relevant chunks highlighted in amber

        Args:
            path (str): File path to write to. e.g. "results.xlsx".
        """
        wb = Workbook()
        wb.remove(wb.active)

        self._write_summary_sheet(wb)
        for experiment in self.experiments:
            self._write_detail_sheet(wb, experiment)

        wb.save(path)
        print(f"Results exported to {path}")

    def to_json(self, path: str) -> None:
        """
        Exports the full experiment results to a JSON file.

        Args:
            path (str): File path to write to. e.g. "results.json".
        """
        # Exclude chunker_object — Python objects are not JSON serialisable.
        # The chunker label (string) is sufficient for JSON export.
        serialisable_experiments = [
            {k: v for k, v in exp.items() if k != "chunker_object"}
            for exp in self.experiments
        ]
        output = {
            "summary": self._summary,
            "experiments": serialisable_experiments,
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        print(f"Results exported to {path}")

    # ── Private helpers ───────────────────────────────────────────────────────

    def _write_summary_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Summary")
        best_label = self.best()["chunker"]

        headers = [
            "Chunker",
            "Best combined (avg)",
            "Best recall (avg)",
            "Best bert (avg)",
            "Avg relevant chunks",
            "Recall threshold",
        ]
        self._write_header_row(ws, headers)

        for i, entry in enumerate(self._summary, start=2):
            is_best = entry["chunker"] == best_label
            fill = _BEST_FILL if is_best else (
                _ALT_FILL if i % 2 == 0 else PatternFill()
            )
            row = [
                entry["chunker"],
                entry["avg_best_combined"],
                entry["avg_best_recall"],
                entry["avg_best_bert"],
                entry["avg_n_relevant"],
                entry["recall_threshold"],
            ]
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.font = _BOLD_FONT if is_best else _NORMAL_FONT
                cell.fill = fill
                cell.border = _THIN_BORDER
                cell.alignment = _WRAP

        ws.column_dimensions["A"].width = 45
        for col in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 20

    def _write_detail_sheet(
        self, wb: Workbook, experiment: dict[str, Any]
    ) -> None:
        label = experiment["chunker"]
        results = experiment["results"]
        sheet_name = label[:31]
        ws = wb.create_sheet(sheet_name)

        n_chunks = results[0]["scores"]["n_retrieved"] if results else 3

        headers = (
            [
                "Question",
                "Expected answer",
                "Best combined",
                "Best recall",
                "Best bert",
                "N relevant",
            ]
            + [f"Chunk {i+1} text" for i in range(n_chunks)]
            + [f"Chunk {i+1} combined" for i in range(n_chunks)]
            + [f"Chunk {i+1} recall" for i in range(n_chunks)]
            + [f"Chunk {i+1} bert" for i in range(n_chunks)]
        )
        self._write_header_row(ws, headers)

        for row_i, result in enumerate(results, start=2):
            scores = result["scores"]
            threshold = scores["recall_threshold"]
            fill = _ALT_FILL if row_i % 2 == 0 else PatternFill()

            base_row = [
                result["question"],
                result["expected_answer"],
                scores["best_combined"],
                scores["best_recall"],
                scores["best_bert"],
                scores["n_relevant"],
            ]
            chunk_texts    = [c["text"] for c in result["retrieved_chunks"]]
            combined_scores = scores["combined_scores"]
            token_recalls   = scores["token_recalls"]
            bert_scores     = scores["bert_scores"]

            full_row = (
                base_row
                + chunk_texts
                + combined_scores
                + token_recalls
                + bert_scores
            )

            for col_i, value in enumerate(full_row, start=1):
                cell = ws.cell(row=row_i, column=col_i, value=value)
                cell.font = _NORMAL_FONT
                cell.border = _THIN_BORDER
                cell.alignment = _WRAP

                # Highlight recall score cells for relevant chunks in amber
                recall_col_start = 6 + n_chunks + n_chunks + 1
                is_recall_col = recall_col_start <= col_i < recall_col_start + n_chunks
                if is_recall_col:
                    chunk_idx = col_i - recall_col_start
                    if chunk_idx < len(token_recalls) and token_recalls[chunk_idx] >= threshold:
                        cell.fill = _RELEVANT_FILL
                    else:
                        cell.fill = fill
                else:
                    cell.fill = fill

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 35
        for col in ["C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 14
        for i in range(n_chunks):
            ws.column_dimensions[get_column_letter(7 + i)].width = 50
        for i in range(n_chunks * 3):
            ws.column_dimensions[get_column_letter(7 + n_chunks + i)].width = 14

        ws.freeze_panes = "A2"

    def _write_header_row(self, ws: Any, headers: list[str]) -> None:
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER

    def _compute_summary(self) -> list[dict[str, Any]]:
        summary = []
        for experiment in self.experiments:
            results = experiment["results"]
            n = len(results)

            avg_best_combined = round(
                sum(r["scores"]["best_combined"] for r in results) / n, 4
            )
            avg_best_recall = round(
                sum(r["scores"]["best_recall"] for r in results) / n, 4
            )
            avg_best_bert = round(
                sum(r["scores"]["best_bert"] for r in results) / n, 4
            )
            avg_n_relevant = round(
                sum(r["scores"]["n_relevant"] for r in results) / n, 1
            )
            recall_threshold = (
                results[0]["scores"]["recall_threshold"] if results else 0.75
            )

            summary.append({
                "chunker":            experiment["chunker"],
                "avg_best_combined":  avg_best_combined,
                "avg_best_recall":    avg_best_recall,
                "avg_best_bert":      avg_best_bert,
                "avg_n_relevant":     avg_n_relevant,
                "avg_overall":        avg_best_combined,
                "recall_threshold":   recall_threshold,
            })

        return sorted(summary, key=lambda e: e["avg_overall"], reverse=True)
